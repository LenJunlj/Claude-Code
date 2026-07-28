"""
CAN Test Tool for FSCM/RSCM - Main GUI Application
Supports PCAN hardware and virtual mode for testing.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import os
from typing import Dict, Optional, Callable

from dbc_parser import DbcParser, DbcDatabase, Message, Signal, classify_messages
from can_comms import CanManager, CanMessage


# ── Constants ──────────────────────────────────────────────────────────
DBC_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', 'CX835_VP2_LDCAN_FSCM_20250211_Fix.dbc')
if not os.path.exists(DBC_PATH):
    DBC_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            '..', 'panel', 'CX835_VP2_LDCAN_FSCM_20250211_Fix.dbc')
if not os.path.exists(DBC_PATH):
    DBC_PATH = r'd:\冷俊\Claude Code\panel\CX835_VP2_LDCAN_FSCM_20250211_Fix.dbc'

FONT_FAMILY = 'Microsoft YaHei'
FONT_SIZE_SMALL = 8
FONT_SIZE_NORMAL = 9
FONT_SIZE_TITLE = 10
COLOR_BG = '#F0F0F0'
COLOR_FRAME_BG = '#E8E8E8'
COLOR_SEND_BG = '#E8F5E9'
COLOR_RECV_BG = '#FFF3E0'
COLOR_ACTIVE = '#4CAF50'
COLOR_WARN = '#FF9800'
COLOR_ERROR = '#F44336'
COLOR_DISABLED = '#CCCCCC'

SEND_TYPE_COLORS = {
    'Cycle': '#1B5E20', 'Event': '#E65100', 'CE': '#6A1B9A',
    'IfActive': '#1565C0', 'CA': '#00838F', 'NoMsgSendType': '#888888'
}

# ── Helpers ────────────────────────────────────────────────────────────

def load_dbc(path: str = DBC_PATH) -> DbcDatabase:
    """Load and parse the DBC file."""
    with open(path, 'r', encoding='gbk') as f:
        text = f.read()
    parser = DbcParser()
    return parser.parse(text)


def build_default_data(msg: Message) -> bytearray:
    """Build default (zeroed) data bytes for a message."""
    return bytearray(msg.dlc)


def encode_message(msg: Message, values: Dict[str, float]) -> bytes:
    """Encode signal values into CAN data bytes."""
    data = bytearray(msg.dlc)
    for sig_name, phys_val in values.items():
        if sig_name in msg.signals:
            msg.signals[sig_name].pack(phys_val, data)
    return bytes(data)


def decode_message(msg: Message, data: bytes) -> Dict[str, float]:
    """Decode CAN data bytes into signal physical values."""
    result = {}
    for sig_name, sig in msg.signals.items():
        if len(data) >= (sig.start_bit + sig.size + 7) // 8:
            result[sig_name] = sig.decode(data)
        else:
            result[sig_name] = 0.0
    return result


EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          '..', 'CX835EX_VP_LDCANFD_FSCM_20251216_Fix.xlsx')


def load_signal_desc_map(excel_path: str = EXCEL_PATH) -> Dict[str, str]:
    """Load Signal Name → Signal Description mapping from Excel matrix.
    Column 7 = signal name, Column 13 = signal description (备注/信号描述).
    """
    import openpyxl
    mapping = {}
    if not os.path.exists(excel_path):
        print(f'Warning: Excel not found at {excel_path}')
        return mapping
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb['FSCM_LDCANFD']
        for row in range(3, ws.max_row + 1):
            sig_name = ws.cell(row=row, column=7).value
            sig_desc = ws.cell(row=row, column=13).value
            if sig_name and str(sig_name).strip():
                sig_name = str(sig_name).strip()
                desc = str(sig_desc).strip() if sig_desc else ''
                if desc:
                    mapping[sig_name] = desc
        wb.close()
        print(f'Loaded {len(mapping)} signal descriptions from Excel')
    except Exception as e:
        print(f'Error reading Excel for signal descriptions: {e}')
    return mapping


def is_analog_signal(sig: Signal) -> bool:
    """Determine if a signal should use analog slider control."""
    if len(sig.value_descriptions) > 0:
        return False
    return sig.size > 1


def is_enum_signal(sig: Signal) -> bool:
    """Determine if a signal uses enum/dropdown control."""
    if len(sig.value_descriptions) > 0:
        return True
    return False


def get_signal_widget_type(sig: Signal) -> str:
    """Classify signal widget type: 'analog', 'enum', or 'binary'."""
    if len(sig.value_descriptions) > 0:
        # If only 2 values (0/1), could be checkbox, but use enum for text
        return 'enum'
    if sig.size == 1:
        return 'binary'
    return 'analog'


# ── Scrollable Frame ───────────────────────────────────────────────────

class ScrollableFrame(ttk.Frame):
    """A frame that is scrollable both vertically and horizontally."""

    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Canvas
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0,
                                bg=COLOR_BG)
        # Vertical scrollbar
        self.v_scrollbar = ttk.Scrollbar(self, orient='vertical',
                                         command=self.canvas.yview)
        # Horizontal scrollbar
        self.h_scrollbar = ttk.Scrollbar(self, orient='horizontal',
                                         command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set,
                              xscrollcommand=self.h_scrollbar.set)

        self.v_scrollbar.grid(row=0, column=1, sticky='ns')
        self.h_scrollbar.grid(row=1, column=0, sticky='ew')
        self.canvas.grid(row=0, column=0, sticky='nsew')

        # Inner frame
        self.inner_frame = ttk.Frame(self.canvas)
        self.inner_frame.bind('<Configure>',
                              lambda e: self.canvas.configure(
                                  scrollregion=self.canvas.bbox('all')))
        self.canvas_window = self.canvas.create_window(
            (0, 0), window=self.inner_frame, anchor='nw',
            tags='inner')

        self.canvas.bind('<Configure>', self._on_canvas_configure)

        # Mouse wheel scrolling
        self.canvas.bind_all('<MouseWheel>', self._on_mousewheel,
                             add='+')

    def _on_canvas_configure(self, event):
        """Set the inner frame width to match canvas."""
        self.canvas.itemconfig('inner', width=event.width)

    def _on_mousewheel(self, event):
        """Scroll on mouse wheel."""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')


# ── Signal Send Widget ─────────────────────────────────────────────────

class AnalogSignalWidget(ttk.Frame):
    """Slider + entry widget for analog signals. Signal name auto-sized."""

    def __init__(self, parent, sig: Signal, on_change: Callable = None,
                 display_name: str = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.sig = sig
        self._on_change = on_change
        self._updating = False

        self.columnconfigure(1, weight=1)

        # Determine precision from factor
        self._digits = max(0, abs(round(str(sig.factor)[::-1].find('.'))) if '.' in str(sig.factor) else 0)
        fmt = f'.{self._digits}f'

        # Label (auto-sized, use display_name if provided)
        label_text = display_name or sig.name
        self.lbl = ttk.Label(self, text=label_text, anchor='w',
                             font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.lbl.grid(row=0, column=0, padx=(0, 0), sticky='w')

        # Value label (auto-sized)
        self.val_lbl = ttk.Label(self, text='0', anchor='e',
                                 font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.val_lbl.grid(row=0, column=2, padx=(0, 0))

        # Unit / precision hint (auto-sized, hidden if no unit)
        unit_text = sig.unit if sig.unit else ''
        hint = f'{unit_text} x{sig.factor}' if unit_text else f'x{sig.factor}'
        self.hint_lbl = ttk.Label(self, text=hint, anchor='w',
                                  font=(FONT_FAMILY, '7'), foreground='#888')
        self.hint_lbl.grid(row=0, column=3, padx=(1, 0))

        # Slider
        from_val = sig.minimum
        to_val = sig.maximum
        if to_val - from_val == 0:
            to_val = 1.0

        self.slider = ttk.Scale(self, from_=from_val, to=to_val,
                                orient='horizontal',
                                command=self._slider_changed,
                                length=70)
        self.slider.grid(row=0, column=1, padx=(0, 0), sticky='ew')
        self.slider.set(0)

    def _slider_changed(self, val):
        if self._updating:
            return
        val = float(val)
        fmt = f'.{self._digits}f'
        self.val_lbl.config(text=f'{val:{fmt}}')
        if self._on_change:
            self._on_change(self.sig.name, val)

    def get_value(self) -> float:
        return float(self.slider.get())

    def set_value(self, val: float, notify: bool = True):
        self._updating = True
        self.slider.set(val)
        fmt = f'.{self._digits}f'
        self.val_lbl.config(text=f'{val:{fmt}}')
        self._updating = False


class EnumSignalWidget(ttk.Frame):
    """Compact combobox dropdown for enum signals. Signal name auto-sized."""

    def __init__(self, parent, sig: Signal, on_change: Callable = None,
                 display_name: str = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.sig = sig
        self._on_change = on_change
        self._updating = False

        self.columnconfigure(0, weight=0)  # label natural width
        self.columnconfigure(1, weight=1)  # combo expands

        # Label (auto-sized, use display_name if provided)
        label_text = display_name or sig.name
        self.lbl = ttk.Label(self, text=label_text, anchor='w',
                             font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.lbl.grid(row=0, column=0, padx=(0, 0), sticky='w')

        # Build value descriptions sorted by value
        sorted_vals = sorted(sig.value_descriptions.items())
        self._val_to_text = {v: t for v, t in sorted_vals}
        self._text_to_val = {t: v for v, t in sorted_vals}

        # Add "reserved" entries for values not in descriptions
        max_val = (1 << sig.size) - 1
        for v in range(max_val + 1):
            if v not in self._val_to_text:
                self._val_to_text[v] = f'r{v}'
                self._text_to_val[f'r{v}'] = v

        values_text = [self._val_to_text[i] for i in sorted(self._val_to_text.keys())]
        combo_width = max(len(str(v)) for v in values_text) + 1

        self.combo = ttk.Combobox(self, values=values_text, width=combo_width,
                                  state='readonly',
                                  font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.combo.grid(row=0, column=1, padx=(0, 0), sticky='ew')
        self.combo.set(values_text[0] if values_text else '')
        self.combo.bind('<<ComboboxSelected>>', self._combo_changed)

    def _combo_changed(self, event=None):
        if self._updating:
            return
        text = self.combo.get()
        if text in self._text_to_val:
            val = self._text_to_val[text]
            if self._on_change:
                self._on_change(self.sig.name, float(val))

    def get_value(self) -> float:
        text = self.combo.get()
        return float(self._text_to_val.get(text, 0))

    def set_value(self, val: float, notify: bool = True):
        self._updating = True
        raw = int(val)
        if raw in self._val_to_text:
            self.combo.set(self._val_to_text[raw])
        self._updating = False


class BinarySignalWidget(ttk.Frame):
    """Compact combobox for binary (0/1) signals. Signal name auto-sized."""

    def __init__(self, parent, sig: Signal, on_change: Callable = None,
                 display_name: str = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.sig = sig
        self._on_change = on_change
        self._updating = False

        self.columnconfigure(0, weight=0)  # label natural width
        self.columnconfigure(2, weight=1)  # unit fills rest

        label_text = display_name or sig.name
        self.lbl = ttk.Label(self, text=label_text, anchor='w',
                             font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.lbl.grid(row=0, column=0, padx=(0, 0), sticky='w')

        self.var = tk.StringVar(value='0')
        combo_values = ['0', '1']
        combo_width = max(len(v) for v in combo_values) + 1
        self.combo = ttk.Combobox(self, textvariable=self.var,
                                  values=combo_values, width=combo_width, state='readonly',
                                  font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.combo.grid(row=0, column=1, padx=(0, 0), sticky='w')
        self.combo.bind('<<ComboboxSelected>>', self._changed)

        # Unit (auto-sized)
        unit_text = sig.unit if sig.unit else ''
        self.unit_lbl = ttk.Label(self, text=unit_text, anchor='w',
                                  font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.unit_lbl.grid(row=0, column=2, padx=(0, 0), sticky='w')

    def _changed(self, event=None):
        if self._updating:
            return
        if self._on_change:
            self._on_change(self.sig.name, float(self.var.get()))

    def get_value(self) -> float:
        return float(self.var.get())

    def set_value(self, val: float, notify: bool = True):
        self._updating = True
        self.var.set(str(int(val)))
        self._updating = False


# ── Send Message Panel ────────────────────────────────────────────────

class SendMessagePanel(ttk.LabelFrame):
    """A message panel with all signal controls for sending.
    Supports: enable/disable toggle, cyclic send, GenSigStartValue init.
    """

    def __init__(self, parent, msg: Message, can_mgr: CanManager = None,
                 app_ref=None, desc_map: Dict[str, str] = None, *args, **kwargs):
        st_color = SEND_TYPE_COLORS.get(msg.send_type_name, '#888')
        title = f'0x{msg.id:03X} {msg.name}  [{msg.send_type_name}]'
        super().__init__(parent, text=title, *args, **kwargs)
        self.msg = msg
        self.can_mgr = can_mgr
        self.app_ref = app_ref
        self._desc_map = desc_map or {}
        self._signal_values: Dict[str, float] = {}
        self._widgets: Dict[str, tk.Widget] = {}
        self._cyclic_active = False

        # Enable/disable control (default: enabled=1)
        self._enabled_var = tk.BooleanVar(value=True)
        self._build_ui()

    def _build_ui(self):
        # Configure 2 columns for horizontal signal layout
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        # Row 0: enable checkbox + message info (span 2 cols)
        self._build_header_row()
        # Signal widgets in 2-column grid (starting row 1)
        self._build_signal_rows()
        # Button row (span 2 cols)
        self._build_button_row()

    def _build_header_row(self):
        """Row 0: Enable checkbox + message info (span 2 cols)."""
        hdr = ttk.Frame(self)
        hdr.grid(row=0, column=0, columnspan=2, sticky='ew', padx=1, pady=(0, 0))
        hdr.columnconfigure(2, weight=1)

        self.chk_enable = ttk.Checkbutton(hdr, text='Enable',
                                          variable=self._enabled_var,
                                          command=self._on_enable_toggled)
        self.chk_enable.grid(row=0, column=0, sticky='w')

        st_color = SEND_TYPE_COLORS.get(self.msg.send_type_name, '#888')
        ttk.Label(hdr,
                  text=f'  {self.msg.send_type_name} {self.msg.cycle_time}ms  DLC={self.msg.dlc}  Tx:{self.msg.transmitter}',
                  font=(FONT_FAMILY, '7'), foreground='#999').grid(
                      row=0, column=2, sticky='e')

    def _build_signal_rows(self):
        """Build signal widgets in 2-column grid (row 1+)."""
        signals_ordered = sorted(self.msg.signals.values(),
                                 key=lambda s: s.start_bit)
        subgroups = self._group_signals_by_seat(signals_ordered)

        row = 1
        for group_name, group_sigs in subgroups:
            if group_name and len(subgroups) > 1:
                lbl = ttk.Label(self, text=group_name,
                                font=(FONT_FAMILY, FONT_SIZE_SMALL, 'bold'),
                                foreground='#555')
                lbl.grid(row=row, column=0, columnspan=2, sticky='w',
                         padx=(2, 0), pady=(0, 0))
                row += 1
            col = 0
            for sig in group_sigs:
                wtype = get_signal_widget_type(sig)
                display_name = self._desc_map.get(sig.name)
                if wtype == 'analog':
                    widget = AnalogSignalWidget(self, sig, self._on_value_change, display_name=display_name)
                elif wtype == 'enum':
                    widget = EnumSignalWidget(self, sig, self._on_value_change, display_name=display_name)
                else:
                    widget = BinarySignalWidget(self, sig, self._on_value_change, display_name=display_name)
                padx_left = 1
                widget.grid(row=row, column=col, sticky='ew', padx=(padx_left, 0), pady=0)
                self._widgets[sig.name] = widget

                # Apply GenSigStartValue if available
                init_val = 0.0
                if sig.gen_sig_start_value is not None:
                    init_val = sig.gen_sig_start_value
                elif sig.name in sig.value_descriptions:
                    init_val = float(min(sig.value_descriptions.keys()))
                widget.set_value(init_val)
                self._signal_values[sig.name] = init_val
                col += 1
                if col >= 2:
                    col = 0
                    row += 1
            if col > 0:
                row += 1  # Advance row for odd count

    def _build_button_row(self):
        """Bottom row: Send Once + Cyclic buttons (span 2 cols)."""
        signals_ordered = sorted(self.msg.signals.values(),
                                 key=lambda s: s.start_bit)
        subgroups = self._group_signals_by_seat(signals_ordered)
        row = 1 + (len(signals_ordered) + 1) // 2 + max(0, len(subgroups) - 1)

        btn_frame = ttk.Frame(self)
        btn_frame.grid(row=row, column=0, columnspan=2, sticky='ew', padx=1, pady=(0, 0))

        self.btn_send = ttk.Button(btn_frame, text='Send Once',
                                   command=self._send_once, width=9)
        self.btn_send.pack(side='left', padx=(0, 1))

        self.btn_cyclic = ttk.Button(btn_frame, text='Cyclic OFF',
                                     command=self._toggle_cyclic, width=9)
        self.btn_cyclic.pack(side='left')

    def _group_signals_by_seat(self, signals):
        """Group signals by seat position for better visual organization."""
        # Detect if this is a seat/mirror message with multiple positions
        prefixes = {}
        ungrouped = []

        for sig in signals:
            # Try to extract seat prefix
            prefix = None
            for pfx in ['LFSeat', 'RFSeat', 'LRSeat', 'RRSeat',
                        'MirrorLeft', 'MirrorRight']:
                if sig.name.startswith(pfx):
                    prefix = pfx
                    break
            if prefix:
                if prefix not in prefixes:
                    prefix_label_map = {
                        'LFSeat': '【Left Front Seat】',
                        'RFSeat': '【Right Front Seat】',
                        'LRSeat': '【Left Rear Seat】',
                        'RRSeat': '【Right Rear Seat】',
                        'MirrorLeft': '【Left Mirror】',
                        'MirrorRight': '【Right Mirror】',
                    }
                    prefixes[prefix] = (prefix_label_map.get(prefix, prefix), [])
                prefixes[prefix][1].append(sig)
            else:
                ungrouped.append(sig)

        result = []
        # Sort by seat position order
        order = ['LFSeat', 'RFSeat', 'LRSeat', 'RRSeat',
                 'MirrorLeft', 'MirrorRight']
        for key in order:
            if key in prefixes:
                result.append(prefixes[key])
        # Add remaining prefixes
        for key in sorted(prefixes.keys()):
            if key not in order:
                result.append(prefixes[key])
        # Add ungrouped (without label if there's only one group)
        if len(prefixes) <= 1:
            result = [('', signals)]
        elif ungrouped:
            result.append(('Other', ungrouped))

        return result

    def is_enabled(self) -> bool:
        return self._enabled_var.get()

    def _on_enable_toggled(self):
        """Handle enable/disable toggle."""
        enabled = self._enabled_var.get()
        state = 'normal' if enabled else 'disabled'
        for widget in self._widgets.values():
            for child in widget.winfo_children():
                try:
                    child.configure(state=state)
                except Exception:
                    pass
        self.btn_send.configure(state=state)
        if not enabled and self._cyclic_active:
            # Auto-stop cyclic when disabled
            self._toggle_cyclic()

    def _on_value_change(self, sig_name: str, value: float):
        self._signal_values[sig_name] = value

    def _send_once(self):
        """Send this message once (only if enabled)."""
        if not self.is_enabled():
            return
        if self.app_ref:
            self.app_ref._send_message(self.msg.id)
        elif self.can_mgr and self.can_mgr.is_connected:
            data = self.get_data_bytes()
            self.can_mgr.send_message(self.msg.id, data)

    def _toggle_cyclic(self):
        """Toggle cyclic sending for this message (only if enabled)."""
        if not self.is_enabled() and not self._cyclic_active:
            return
        if self.app_ref:
            if not self._cyclic_active:
                self.app_ref._start_cyclic_id(self.msg.id)
                self._cyclic_active = True
                self.btn_cyclic.config(text='Cyclic ON',
                                       style='Cyclic.TButton')
            else:
                self.app_ref._stop_cyclic_id(self.msg.id)
                self._cyclic_active = False
                self.btn_cyclic.config(text='Cyclic OFF',
                                       style='TButton')

    def set_cyclic_state(self, active: bool):
        """Update cyclic state from external."""
        self._cyclic_active = active
        try:
            if active:
                self.btn_cyclic.config(text='Cyclic ON',
                                       style='Cyclic.TButton')
            else:
                self.btn_cyclic.config(text='Cyclic OFF',
                                       style='TButton')
        except Exception:
            pass

    def get_signal_values(self) -> Dict[str, float]:
        return dict(self._signal_values)

    def get_data_bytes(self) -> bytes:
        if not self.is_enabled():
            return b''
        return encode_message(self.msg, self._signal_values)


class SendPanel(ttk.Frame):
    """Left panel: All send controls grouped by message."""

    # Category definitions: (label, color, message_ids)
    CATEGORIES = [
        ('Seat Control - Front', '#1B5E20',
         [710, 711]),  # CCU_FSCMCTRL, CCU_RSCMCTRL
        ('Mirror Control', '#1565C0',
         [712]),  # CCU_MirrorCTRL
        ('Vehicle Status', '#6A1B9A',
         [1412, 1264, 1265, 1296, 545, 291]),
        ('Body & Safety', '#E65100',
         [784, 801, 848, 1448, 706]),
    ]

    def __init__(self, parent, to_fscm: Dict[int, Message],
                 can_mgr: CanManager, app_ref=None,
                 desc_map: Dict[str, str] = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.can_mgr = can_mgr
        self.to_fscm = to_fscm
        self.app_ref = app_ref
        self._desc_map = desc_map or {}
        self._msg_panels: Dict[int, SendMessagePanel] = {}

        # Build category map
        self._cat_map: Dict[int, str] = {}
        for cat_name, color, mids in self.CATEGORIES:
            for mid in mids:
                if mid in to_fscm:
                    self._cat_map[mid] = cat_name

        # Header
        self._build_header()

        # Scrollable area for message panels
        self.scroll = ScrollableFrame(self)
        self.scroll.grid(row=1, column=0, sticky='nsew')
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.inner = self.scroll.inner_frame

        # Build message panels in category order
        self._build_by_category()

    def _build_header(self):
        header_frame = ttk.Frame(self)
        header_frame.grid(row=0, column=0, sticky='ew', padx=0, pady=0)
        self.grid_columnconfigure(0, weight=1)

        ttk.Label(header_frame, text='▶ TRANSMIT (to FSCM/RSCM)',
                  font=(FONT_FAMILY, FONT_SIZE_TITLE, 'bold'),
                  foreground='#2E7D32').pack(side='left')

    def _build_by_category(self):
        """Build message panels organized by category."""
        row = 0
        processed = set()

        for cat_name, color, mids in self.CATEGORIES:
            msg_ids_in_cat = [mid for mid in mids if mid in self.to_fscm
                              and mid not in processed]
            if not msg_ids_in_cat:
                continue

            # Category header
            cat_frame = ttk.Frame(self.inner)
            cat_frame.grid(row=row, column=0, sticky='ew', padx=1, pady=(1, 0))
            cat_lbl = tk.Label(cat_frame,
                               text=f'── {cat_name} ──',
                               font=(FONT_FAMILY, FONT_SIZE_NORMAL, 'bold'),
                               fg=color, bg=COLOR_FRAME_BG,
                               anchor='w')
            cat_lbl.pack(fill='x')
            row += 1

            for msg_id in msg_ids_in_cat:
                msg = self.to_fscm.get(msg_id)
                if msg is None:
                    continue
                panel = SendMessagePanel(self.inner, msg,
                                         can_mgr=self.can_mgr,
                                         app_ref=self.app_ref,
                                         desc_map=self._desc_map)
                panel.grid(row=row, column=0, sticky='ew', padx=1, pady=0)
                self._msg_panels[msg_id] = panel
                processed.add(msg_id)
                row += 1

        # Remaining messages not in predefined categories
        remaining = [mid for mid in sorted(self.to_fscm.keys())
                     if mid not in processed]
        if remaining:
            cat_frame = ttk.Frame(self.inner)
            cat_frame.grid(row=row, column=0, sticky='ew', padx=1, pady=(1, 0))
            tk.Label(cat_frame,
                     text='── Other Signals ──',
                     font=(FONT_FAMILY, FONT_SIZE_NORMAL, 'bold'),
                     fg='#888888', bg=COLOR_FRAME_BG,
                     anchor='w').pack(fill='x')
            row += 1
            for msg_id in remaining:
                msg = self.to_fscm.get(msg_id)
                panel = SendMessagePanel(self.inner, msg,
                                         can_mgr=self.can_mgr,
                                         app_ref=self.app_ref,
                                         desc_map=self._desc_map)
                panel.grid(row=row, column=0, sticky='ew', padx=1, pady=0)
                self._msg_panels[msg_id] = panel
                row += 1

    def is_msg_enabled(self, msg_id: int) -> bool:
        panel = self._msg_panels.get(msg_id)
        return panel is not None and panel.is_enabled()

    def get_msg_data(self, msg_id: int) -> Optional[bytes]:
        panel = self._msg_panels.get(msg_id)
        if panel and panel.is_enabled():
            return panel.get_data_bytes()
        return None

    def set_cyclic_state(self, msg_id: int, active: bool):
        panel = self._msg_panels.get(msg_id)
        if panel:
            panel.set_cyclic_state(active)


# ── Receive Display Panel ─────────────────────────────────────────────

class RecvSignalDisplay(ttk.Frame):
    """Display widget for a single received signal value. Signal name auto-sized."""

    def __init__(self, parent, sig: Signal, display_name: str = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.sig = sig
        self._flash_job = None

        self.columnconfigure(0, weight=0)  # label natural width
        self.columnconfigure(1, weight=1)  # value expands

        # Label (auto-sized, use display_name if provided)
        label_text = display_name or sig.name
        self.lbl = ttk.Label(self, text=label_text, anchor='w',
                             font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.lbl.grid(row=0, column=0, padx=(0, 0), sticky='w')

        # Value display (auto-sized)
        self.val_lbl = ttk.Label(self, text='--', anchor='w',
                                 font=(FONT_FAMILY, FONT_SIZE_SMALL, 'bold'),
                                 foreground='#333333')
        self.val_lbl.grid(row=0, column=1, padx=(0, 0), sticky='w')

        # Unit (auto-sized)
        unit_text = sig.unit if sig.unit else ''
        self.unit_lbl = ttk.Label(self, text=unit_text, anchor='w',
                                  font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.unit_lbl.grid(row=0, column=2, padx=(0, 0), sticky='w')

    def update_value(self, phys_val: float, raw: int):
        """Update the displayed value and flash briefly."""
        sig = self.sig
        if raw in sig.value_descriptions:
            text = f'{sig.value_descriptions[raw]} ({raw})'
        elif sig.factor != 1 or sig.offset != 0:
            text = f'{phys_val:.1f}'
            # Show raw in parentheses for analog
            text = f'{phys_val:.2f} ({raw})'
        else:
            text = f'{int(phys_val)} ({raw})'

        self.val_lbl.config(text=text, foreground='#0066CC')

        # Flash effect
        if self._flash_job:
            self.after_cancel(self._flash_job)
        self._flash_job = self.after(200, lambda: self.val_lbl.config(
            foreground='#333333') if self.winfo_exists() else None)


class RecvMessagePanel(ttk.LabelFrame):
    """A display panel for a received message with all signal values."""

    def __init__(self, parent, msg: Message, desc_map: Dict[str, str] = None, *args, **kwargs):
        title = f'0x{msg.id:03X} {msg.name} (Tx: {msg.transmitter})'
        super().__init__(parent, text=title, *args, **kwargs)
        self.msg = msg
        self._desc_map = desc_map or {}
        self._displays: Dict[str, RecvSignalDisplay] = {}
        self._last_values: Dict[str, tuple] = {}

        self._build_ui()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=1)
        signals_ordered = sorted(self.msg.signals.values(),
                                 key=lambda s: s.start_bit)
        row = 0
        col = 0
        for sig in signals_ordered:
            disp = RecvSignalDisplay(self, sig, display_name=self._desc_map.get(sig.name))
            disp.grid(row=row, column=col, sticky='ew', padx=1, pady=0)
            self._displays[sig.name] = disp
            col += 1
            if col >= 2:
                col = 0
                row += 1
        if col > 0:
            row += 1

    def update_from_data(self, data: bytes):
        """Update all signal displays from raw CAN data."""
        for sig_name, sig in self.msg.signals.items():
            if len(data) * 8 < sig.start_bit + sig.size:
                continue
            try:
                raw = sig._extract_raw(data)
                phys = sig.decode(data)
                # Check if value changed
                key = (raw,)  # just compare raw
                if key != self._last_values.get(sig_name):
                    self._last_values[sig_name] = key
                    disp = self._displays.get(sig_name)
                    if disp:
                        disp.update_value(phys, raw)
            except Exception:
                pass

    def get_timestamp_label(self) -> str:
        return time.strftime('%H:%M:%S.%f')[:-3]


class RecvPanel(ttk.Frame):
    """Right panel: Received message displays for a single function group.
    Dynamically creates panels for any DBC-known message ID on the bus.
    """

    def __init__(self, parent, from_fscm: Dict[int, Message],
                 can_mgr: CanManager, db_messages: Dict[int, Message] = None,
                 desc_map: Dict[str, str] = None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.can_mgr = can_mgr
        self.from_fscm = from_fscm
        self._db_messages = db_messages or {}
        self._desc_map = desc_map or {}
        self._msg_displays: Dict[int, RecvMessagePanel] = {}
        self._panel_row_counter = 0

        # Header
        self._build_header()

        # Scrollable area
        self.scroll = ScrollableFrame(self)
        self.scroll.grid(row=1, column=0, sticky='nsew')
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.inner = self.scroll.inner_frame

        # Pre-build panels for known messages
        for msg_id in sorted(from_fscm.keys()):
            msg = from_fscm.get(msg_id)
            if msg is None:
                continue
            self._add_panel(msg)

    def _build_header(self):
        header = ttk.Label(self, text='▶ RECEIVE',
                           font=(FONT_FAMILY, FONT_SIZE_TITLE, 'bold'),
                           foreground='#E65100',
                           background=COLOR_RECV_BG)
        header.grid(row=0, column=0, sticky='ew', padx=0, pady=0)
        self.grid_columnconfigure(0, weight=1)

    def _add_panel(self, msg: Message):
        """Create and add a RecvMessagePanel."""
        if msg.id in self._msg_displays:
            return
        panel = RecvMessagePanel(self.inner, msg, desc_map=self._desc_map)
        panel.grid(row=self._panel_row_counter, column=0, sticky='ew',
                   padx=1, pady=0)
        self._msg_displays[msg.id] = panel
        self._panel_row_counter += 1

    def update_message(self, msg_id: int, data: bytes):
        """Update display for a received message.
        Auto-creates a panel if the message is in the DBC but not yet displayed.
        """
        panel = self._msg_displays.get(msg_id)
        if panel:
            panel.update_from_data(data)
        elif msg_id in self._db_messages:
            msg = self._db_messages[msg_id]
            self._add_panel(msg)
            self._msg_displays[msg_id].update_from_data(data)


# ── Main Application ───────────────────────────────────────────────────

class CanTestTool(tk.Tk):
    """Main application window."""

    def __init__(self, db: DbcDatabase):
        super().__init__()
        self.db = db
        self.can_mgr = CanManager()
        self.to_fscm, self.from_fscm = classify_messages(db)
        self.signal_desc_map = load_signal_desc_map()

        self.title('CAN Test Tool - FSCM/RSCM')
        self.geometry('1400x850')
        self.minsize(1000, 600)
        self.configure(bg=COLOR_BG)

        # Styles
        style = ttk.Style()
        style.configure('Cyclic.TButton', foreground='#4CAF50', font=('', 9, 'bold'))

        # Set icon if available
        try:
            self.iconbitmap(default='')
        except Exception:
            pass

        # Build UI
        self._build_menu()
        self._build_toolbar()
        self._build_main_area()
        self._build_statusbar()

        # Track cyclic timers
        self._cyclic_timers: Dict[int, threading.Timer] = {}
        self._cyclic_running = False

        # RX callback
        self.can_mgr.register_rx_all_callback(self._on_rx_message)

        # Protocol
        self.protocol('WM_DELETE_WINDOW', self._on_close)

        # Auto-scan for available devices after UI is built
        self.after(200, self._refresh_devices)

    # ── Menu ───────────────────────────────────────────────────────

    def _build_menu(self):
        menubar = tk.Menu(self, font=(FONT_FAMILY, FONT_SIZE_NORMAL))

        # File menu
        file_menu = tk.Menu(menubar, tearoff=0,
                            font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        file_menu.add_command(label='Load DBC...', command=self._load_dbc,
                              accelerator='Ctrl+O')
        file_menu.add_separator()
        file_menu.add_command(label='Exit', command=self._on_close,
                              accelerator='Ctrl+Q')
        menubar.add_cascade(label='File', menu=file_menu)

        # CAN menu
        can_menu = tk.Menu(menubar, tearoff=0,
                           font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        can_menu.add_command(label='Connect (selected channel)',
                             command=self._connect, accelerator='Ctrl+C')
        can_menu.add_command(label='Connect Virtual',
                             command=self._connect_virtual)
        can_menu.add_command(label='Disconnect', command=self._disconnect,
                             accelerator='Ctrl+D')
        can_menu.add_separator()
        can_menu.add_command(label='Start Cyclic Send All',
                             command=self._start_all_cyclic)
        can_menu.add_command(label='Stop Cyclic Send All',
                             command=self._stop_all_cyclic)
        menubar.add_cascade(label='CAN', menu=can_menu)

        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0,
                            font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        help_menu.add_command(label='About', command=self._show_about)
        menubar.add_cascade(label='Help', menu=help_menu)

        self.config(menu=menubar)

        # Keyboard shortcuts
        self.bind('<Control-o>', lambda e: self._load_dbc())
        self.bind('<Control-q>', lambda e: self._on_close())
        self.bind('<Control-c>', lambda e: self._connect())
        self.bind('<Control-d>', lambda e: self._disconnect())

    # ── Toolbar ────────────────────────────────────────────────────

    def _build_toolbar(self):
        toolbar = ttk.Frame(self, padding=(1, 0))
        toolbar.grid(row=0, column=0, sticky='ew', padx=0, pady=(0, 0))
        self.grid_columnconfigure(0, weight=1)

        # ── Row 0: CAN Configuration ──

        # CAN channel selection
        ttk.Label(toolbar, text='Ch:',
                  font=(FONT_FAMILY, FONT_SIZE_NORMAL)).grid(
                      row=0, column=0, padx=(0, 1))
        self.channel_var = tk.StringVar(value='PCAN_USBBUS1')
        self.channel_combo = ttk.Combobox(toolbar, textvariable=self.channel_var,
                                     values=[
                                         'PCAN_USBBUS1', 'PCAN_USBBUS2',
                                         'PCAN_ISABUS1', 'PCAN_ISABUS2',
                                         'virtual'
                                     ],
                                     state='normal', width=14,
                                     font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.channel_combo.grid(row=0, column=1, padx=(0, 1))

        # Refresh devices button
        self.btn_refresh = ttk.Button(toolbar, text='⟳',
                                      command=self._refresh_devices,
                                      width=3)
        self.btn_refresh.grid(row=0, column=2, padx=(0, 2))

        # Mode: CAN / CAN FD
        ttk.Label(toolbar, text='Mode:',
                  font=(FONT_FAMILY, FONT_SIZE_NORMAL)).grid(
                      row=0, column=3, padx=(0, 2))
        self.mode_var = tk.StringVar(value='CAN')
        self.mode_combo = ttk.Combobox(toolbar, textvariable=self.mode_var,
                                     values=['CAN', 'CAN FD'],
                                     state='readonly', width=6,
                                     font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.mode_combo.grid(row=0, column=4, padx=(0, 2))
        self.mode_combo.bind('<<ComboboxSelected>>', self._on_mode_changed)

        # Baud rate (arbitration, kbaud)
        ttk.Label(toolbar, text='Baud:',
                  font=(FONT_FAMILY, FONT_SIZE_NORMAL)).grid(
                      row=0, column=5, padx=(0, 2))
        self.baud_var = tk.StringVar(value='500')
        self.baud_combo = ttk.Combobox(toolbar, textvariable=self.baud_var,
                                     values=['125', '250', '500', '1000'],
                                     state='normal', width=7,
                                     font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.baud_combo.grid(row=0, column=6, padx=(0, 2))

        # Data rate (data phase, kbaud, only for CAN FD)
        ttk.Label(toolbar, text='Data:',
                  font=(FONT_FAMILY, FONT_SIZE_NORMAL)).grid(
                      row=0, column=7, padx=(0, 2))
        self.data_var = tk.StringVar(value='2000')
        self.data_combo = ttk.Combobox(toolbar, textvariable=self.data_var,
                                     values=['500', '1000', '2000', '4000', '5000', '8000'],
                                     state='disabled', width=7,
                                     font=(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.data_combo.grid(row=0, column=8, padx=(0, 2))

        # ── Row 1: Action buttons ──

        # Connect/Disconnect buttons
        self.btn_connect = ttk.Button(toolbar, text='Connect',
                                      command=self._connect,
                                      width=10)
        self.btn_connect.grid(row=1, column=0, padx=(0, 2), pady=(1, 0))

        self.btn_disconnect = ttk.Button(toolbar, text='Disconnect',
                                         command=self._disconnect,
                                         width=10, state='disabled')
        self.btn_disconnect.grid(row=1, column=1, padx=(0, 5), pady=(1, 0))

        # Separator
        sep = ttk.Separator(toolbar, orient='vertical')
        sep.grid(row=1, column=2, sticky='ns', padx=2, pady=1)

        # Cyclic send control
        self.cyclic_var = tk.BooleanVar(value=False)
        self.chk_cyclic = ttk.Checkbutton(
            toolbar, text='Cyclic Send',
            variable=self.cyclic_var,
            command=self._toggle_cyclic_all,
            state='disabled')
        self.chk_cyclic.grid(row=1, column=3, padx=(0, 5), pady=(1, 0))

        # Send Once button
        self.btn_send_once = ttk.Button(
            toolbar, text='Send All Once',
            command=self._send_all_once,
            width=14, state='disabled')
        self.btn_send_once.grid(row=1, column=4, padx=(0, 2), pady=(1, 0))

    def _on_mode_changed(self, event=None):
        """Enable/disable data rate combo based on mode selection."""
        if self.mode_var.get() == 'CAN FD':
            self.data_combo.config(state='normal')
        else:
            self.data_combo.config(state='disabled')

    # ── Main Area (single PanedWindow) ───────────────────────────

    def _build_main_area(self):
        """Build a single PanedWindow with send (left) and recv (right) panels."""
        self.paned = ttk.PanedWindow(self, orient='horizontal')
        self.paned.grid(row=1, column=0, sticky='nsew', padx=0, pady=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Left: Send panel (all to_fscm messages)
        left_frame = ttk.Frame(self.paned)
        self.send_panel = SendPanel(left_frame, self.to_fscm, self.can_mgr,
                                    app_ref=self,
                                    desc_map=self.signal_desc_map)
        self.send_panel.pack(fill='both', expand=True)
        self.paned.add(left_frame, weight=1)

        # Right: Recv panel (all from_fscm messages + full db_messages for dynamic creation)
        right_frame = ttk.Frame(self.paned)
        self.recv_panel = RecvPanel(right_frame, self.from_fscm, self.can_mgr,
                                    db_messages=self.db.messages,
                                    desc_map=self.signal_desc_map)
        self.recv_panel.pack(fill='both', expand=True)
        self.paned.add(right_frame, weight=1)

    # ── Status Bar ─────────────────────────────────────────────────

    def _build_statusbar(self):
        self.statusbar = ttk.Frame(self, relief='sunken', padding=(1, 0))
        self.statusbar.grid(row=2, column=0, sticky='ew', padx=0, pady=(0, 0))

        self.status_label = ttk.Label(
            self.statusbar, text='Status: Disconnected',
            font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.status_label.pack(side='left', padx=2)

        self.rx_count_label = ttk.Label(
            self.statusbar, text='RX: 0',
            font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.rx_count_label.pack(side='right', padx=5)

        self.tx_count_label = ttk.Label(
            self.statusbar, text='TX: 0',
            font=(FONT_FAMILY, FONT_SIZE_SMALL))
        self.tx_count_label.pack(side='right', padx=5)

        self._rx_count = 0
        self._tx_count = 0

    # ── CAN Operations ─────────────────────────────────────────────

    def _refresh_devices(self):
        """Scan for available CAN devices and update the dropdown."""
        try:
            from can_comms import CanManager
            devices = CanManager.list_available_devices()
            current = self.channel_var.get()
            self.channel_combo['values'] = devices
            if current not in devices:
                self.channel_var.set(devices[0] if devices else 'PCAN_USBBUS1')
            self.status_label.config(text=f'Scan: found {len(devices)} device(s)',
                                     foreground='blue')
            self.after(2000, lambda: self.status_label.config(
                text=f'Status: {"Connected" if self.can_mgr.is_connected else "Disconnected"}',
                foreground='green' if self.can_mgr.is_connected else 'black'))
        except Exception as e:
            print(f'Device scan error: {e}')

    def _get_can_config(self) -> tuple:
        """Parse CAN config: (baud_int, fd_mode, data_rate_int)."""
        try:
            baud = int(self.baud_var.get()) * 1000
        except (ValueError, TypeError):
            baud = 500000
        fd_mode = self.mode_var.get() == 'CAN FD'
        try:
            data_rate = int(self.data_var.get()) * 1000
        except (ValueError, TypeError):
            data_rate = 2000000
        return baud, fd_mode, data_rate

    def _connect(self):
        channel = self.channel_var.get()
        if channel == 'virtual':
            self._connect_virtual()
        else:
            self._connect_hardware(channel)

    def _connect_hardware(self, channel: str):
        """Connect to PCAN or Vector hardware with the selected CAN config."""
        baud, fd_mode, data_rate = self._get_can_config()
        # Numeric channels or non-PCAN names → Vector/CANoe interface
        is_vector = channel.isdigit() or channel == 'CANoe'
        if is_vector:
            ok = self.can_mgr.connect_canoe(channel, baud, fd_mode, data_rate)
        else:
            ok = self.can_mgr.connect_pcan(channel, baud, fd_mode, data_rate)
        if ok:
            self._on_connected()
        else:
            msg = (f'Failed to connect to {channel} @ {baud//1000} kbaud.\n'
                   'Check hardware is connected and driver installed.')
            messagebox.showerror('Connection Error', msg)

    def _connect_virtual(self):
        if self.can_mgr.connect_virtual():
            self._on_connected()
            messagebox.showinfo('Virtual Mode',
                                'Connected in virtual mode.\n'
                                'Messages will loop back for testing.')
        else:
            messagebox.showerror('Error', 'Failed to start virtual mode.')

    def _on_connected(self):
        self.btn_connect.config(state='disabled')
        self.btn_disconnect.config(state='normal')
        self.btn_send_once.config(state='normal')
        self.chk_cyclic.config(state='normal')
        self.status_label.config(text=f'Status: Connected ({self.can_mgr.interface_name})',
                                 foreground='green')
        # Update UI with connection state
        self._update_can_state(True)

    def _disconnect(self):
        self.cyclic_var.set(False)
        self._stop_all_cyclic()
        self.can_mgr.disconnect()
        self.on_disconnect()
        self.btn_connect.config(state='normal')
        self.btn_disconnect.config(state='disabled')
        self.btn_send_once.config(state='disabled')
        self.chk_cyclic.config(state='disabled')
        self.status_label.config(text='Status: Disconnected', foreground='black')
        self._update_can_state(False)

    def _update_can_state(self, connected: bool):
        """Update UI for connection state changes."""
        state = 'normal' if connected else 'disabled'
        # Enable/disable per-message send buttons if any

    def _send_all_once(self):
        """Send all messages once."""
        if not self.can_mgr.is_connected:
            return
        sent = set()
        for msg_id in self.to_fscm:
            if msg_id in sent:
                continue
            data = self.send_panel.get_msg_data(msg_id)
            if data:
                self.can_mgr.send_message(msg_id, data)
                self._tx_count += 1
                sent.add(msg_id)
        self.tx_count_label.config(text=f'TX: {self._tx_count}')

    def _toggle_cyclic_all(self):
        """Toggle cyclic send for all messages."""
        if self.cyclic_var.get():
            self._start_all_cyclic()
        else:
            self._stop_all_cyclic()

    def _start_all_cyclic(self):
        """Start cyclic sending for all messages."""
        if not self.can_mgr.is_connected:
            return
        self._cyclic_running = True
        sent = set()
        for msg_id, msg in self.to_fscm.items():
            if msg_id in sent:
                continue
            sent.add(msg_id)
            interval = msg.cycle_time if msg.cycle_time > 0 else 100
            self.can_mgr.start_cyclic_send(
                msg_id,
                lambda mid=msg_id: self.send_panel.get_msg_data(mid) or b'',
                interval
            )
        self.cyclic_var.set(True)
        self.status_label.config(text=f'Status: Connected - Cyclic Sending ({self.can_mgr.interface_name})',
                                 foreground='green')

    def _stop_all_cyclic(self):
        """Stop all cyclic sending."""
        self._cyclic_running = False
        self.can_mgr.stop_all_cyclic()
        self.cyclic_var.set(False)
        self.status_label.config(text=f'Status: Connected ({self.can_mgr.interface_name})',
                                 foreground='green')

    def _send_message(self, msg_id: int):
        """Send a single message."""
        if not self.can_mgr.is_connected:
            return
        data = self.send_panel.get_msg_data(msg_id)
        if data:
            self.can_mgr.send_message(msg_id, data)
            self._tx_count += 1
            self.tx_count_label.config(text=f'TX: {self._tx_count}')

    def _start_cyclic_id(self, msg_id: int):
        """Start cyclic sending for a specific message."""
        if not self.can_mgr.is_connected:
            return
        msg = self.to_fscm.get(msg_id)
        if not msg:
            return
        interval = msg.cycle_time if msg.cycle_time > 0 else 100
        self.can_mgr.start_cyclic_send(
            msg_id,
            lambda mid=msg_id: self.send_panel.get_msg_data(mid) or b'',
            interval
        )

    def _stop_cyclic_id(self, msg_id: int):
        """Stop cyclic sending for a specific message."""
        self.can_mgr.stop_cyclic_send(msg_id)
        self.send_panel.set_cyclic_state(msg_id, False)

    def on_disconnect(self):
        """Called when disconnected - reset all cyclic states."""
        self._stop_all_cyclic()
        for msg_id in self.to_fscm:
            self.send_panel.set_cyclic_state(msg_id, False)

    # ── RX Handling ────────────────────────────────────────────────

    def _on_rx_message(self, msg: CanMessage):
        """Handle received CAN message. Update the recv panel."""
        self._rx_count += 1
        self.after(0, lambda: self.rx_count_label.config(
            text=f'RX: {self._rx_count}'))

        if msg.arb_id in self.db.messages:
            self.after(0, lambda mid=msg.arb_id, d=msg.data:
                       self.recv_panel.update_message(mid, d))

    # ── File Operations ────────────────────────────────────────────

    def _load_dbc(self):
        """Load a different DBC file."""
        filepath = filedialog.askopenfilename(
            title='Select DBC File',
            filetypes=[('DBC Files', '*.dbc'), ('All Files', '*.*')])
        if not filepath:
            return
        try:
            new_db = load_dbc(filepath)
            self.db = new_db
            self.to_fscm, self.from_fscm = classify_messages(new_db)

            # Rebuild main area
            self.paned.destroy()
            self._build_main_area()

            # Re-register RX callbacks
            # (keep existing manager)

            self.status_label.config(text=f'Loaded: {os.path.basename(filepath)}',
                                     foreground='blue')
        except Exception as e:
            messagebox.showerror('Error', f'Failed to load DBC: {e}')

    # ── Misc ───────────────────────────────────────────────────────

    def _show_about(self):
        messagebox.showinfo(
            'About CAN Test Tool',
            'CAN Test Tool for FSCM/RSCM\n\n'
            'Based on DBC: CX835_VP2_LDCAN_FSCM_20250211_Fix.dbc\n\n'
            'Supports PCAN USB hardware and virtual mode.\n'
            'Signal grouping by message for easy testing.'
        )

    def _on_close(self):
        """Clean up on window close."""
        self._stop_all_cyclic()
        self.can_mgr.disconnect()
        self.destroy()


# ── Entry Point ────────────────────────────────────────────────────────

def main():
    # Load DBC
    print(f'Loading DBC: {DBC_PATH}')
    db = load_dbc(DBC_PATH)
    print(f'Loaded {len(db.messages)} messages, {len(db.nodes)} nodes')

    # Start application
    app = CanTestTool(db)
    app.mainloop()


if __name__ == '__main__':
    main()
