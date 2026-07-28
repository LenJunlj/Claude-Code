"""
Generate a DBC file from the Excel communication matrix.
The Excel (2025-12-16) is the authoritative source vs the old DBC (2025-02-11).
All signals are Motorola MSB.
"""
import openpyxl
import re
import os

EXCEL_PATH = r'd:\冷俊\Claude Code\panel\CX835EX_VP_LDCANFD_FSCM_20251216_Fix.xlsx'
OUTPUT_DBC = r'd:\冷俊\Claude Code\panel\CX835_VP2_LDCAN_FSCM_20250211_Fix.dbc'


def parse_val_desc(text):
    """Parse value description text like '0x0:P\\r\\n0x1:L_Reserved\\r\\n...' into {0: 'P', 1: 'L_Reserved'...}"""
    result = {}
    if not text:
        return result
    # Split on common delimiters
    parts = re.split(r'[\\r\\n\\n]+', text)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Try "0x0:P" or "0x0 P" or "0x0:P value" format
        m = re.match(r'(0x[0-9A-Fa-f]+)\s*[:=]?\s*(.*)', part)
        if m:
            try:
                val = int(m.group(1), 16)
                desc = m.group(2).strip()
                result[val] = desc
            except ValueError:
                pass
    return result


def build_dbc(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['FSCM_LDCANFD']

    # Collect node names from column headers (col 34+)
    nodes_set = set()
    for col in range(34, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val and str(val).strip():
            nodes_set.add(str(val).strip())

    # Ensure standard nodes
    for n in ['FSCM', 'CCU', 'LH', 'SOA', 'SSCM', 'ZCUL']:
        if n not in nodes_set:
            nodes_set.add(n)

    # Map send types
    SEND_TYPE_MAP = {
        'Cycle': 0, 'NoMsgSendType': 1, 'IfActive': 2,
        'Event': 3, 'CA': 4, 'CE': 5,
    }

    # First pass: collect all messages and signals
    messages = {}  # msg_id_int -> dict
    current_msg_id = None

    for row_idx in range(3, ws.max_row + 1):
        msg_name = ws.cell(row=row_idx, column=1).value
        if msg_name and str(msg_name).strip():
            msg_name = str(msg_name).strip()
            msg_id_str = ws.cell(row=row_idx, column=3).value
            msg_type = ws.cell(row=row_idx, column=2).value
            msg_send_type = ws.cell(row=row_idx, column=4).value
            msg_cycle = ws.cell(row=row_idx, column=5).value
            msg_len = ws.cell(row=row_idx, column=6).value
            # Determine transmitter from node columns (col 34+) - first "Tx" found
            transmitter = 'FSCM'
            for col in range(34, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val and str(cell_val).strip() == 'Tx':
                    node_name = ws.cell(row=1, column=col).value
                    if node_name:
                        transmitter = str(node_name).strip()
                    break
                elif cell_val and str(cell_val).strip() == 'Rx':
                    continue

            msg_id_str = str(msg_id_str).strip() if msg_id_str else ''
            msg_id_int = int(msg_id_str, 16) if msg_id_str.startswith('0x') else 0
            msg_len = int(msg_len) if msg_len else 8
            msg_cycle = int(msg_cycle) if msg_cycle else 0
            send_type = int(SEND_TYPE_MAP.get(str(msg_send_type).strip(), 0)) if msg_send_type else 0

            # Detect NM messages by name pattern
            is_nm = '_NM_' in msg_name.upper() or msg_name.upper().endswith('_NM')

            current_msg_id = msg_id_int
            messages[msg_id_int] = {
                'name': msg_name,
                'dlc': msg_len,
                'transmitter': transmitter,
                'cycle_time': msg_cycle,
                'send_type': send_type,
                'is_nm': is_nm,
                'signals': [],
            }

        sig_name = ws.cell(row=row_idx, column=7).value
        if sig_name and str(sig_name).strip() and current_msg_id is not None:
            sig_name = str(sig_name).strip()
            byte_order_str = str(ws.cell(row=row_idx, column=14).value or '')
            start_bit = ws.cell(row=row_idx, column=16).value
            sig_len = ws.cell(row=row_idx, column=18).value
            data_type_str = str(ws.cell(row=row_idx, column=19).value or '')
            resolution = ws.cell(row=row_idx, column=20).value
            offset = ws.cell(row=row_idx, column=21).value
            min_phys = ws.cell(row=row_idx, column=22).value
            max_phys = ws.cell(row=row_idx, column=23).value
            unit = ws.cell(row=row_idx, column=29).value
            val_desc_str = ws.cell(row=row_idx, column=30).value
            sig_desc = ws.cell(row=row_idx, column=13).value
            init_hex = ws.cell(row=row_idx, column=26).value
            invalid_hex = ws.cell(row=row_idx, column=27).value

            # Determine byte order: 1 for Motorola, 0 for Intel
            if 'Motorola' in byte_order_str:
                byte_order = 1
            elif 'Intel' in byte_order_str:
                byte_order = 0
            else:
                byte_order = 1  # Default to Motorola as per ECU spec

            # Determine signed/unsigned
            signed = 'signed' in data_type_str.lower()

            # Parse factor and offset
            factor = float(resolution) if resolution else 1.0
            offset_val = float(offset) if offset else 0.0

            # Parse min/max
            min_v = float(min_phys) if min_phys else 0.0
            max_v = float(max_phys) if max_phys else 0.0

            # Determine receivers from node columns (col 34+)
            receivers = []
            for col in range(34, ws.max_column + 1):
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val and str(cell_val).strip() == 'Rx':
                    node_name = ws.cell(row=1, column=col).value
                    if node_name:
                        receivers.append(str(node_name).strip())

            # Parse value descriptions
            val_desc = parse_val_desc(val_desc_str)

            # Parse initial value
            init_raw = None
            if init_hex:
                init_hex_clean = init_hex.strip()
                if init_hex_clean.startswith('0x'):
                    try:
                        init_raw = int(init_hex_clean, 16)
                    except ValueError:
                        pass

            # Parse invalid value
            invalid_raw = None
            if invalid_hex:
                invalid_hex_clean = invalid_hex.strip()
                if invalid_hex_clean.startswith('0x'):
                    try:
                        invalid_raw = int(invalid_hex_clean, 16)
                    except ValueError:
                        pass

            start_bit = int(start_bit) if start_bit else 0
            sig_len = int(sig_len) if sig_len else 1

            messages[current_msg_id]['signals'].append({
                'name': sig_name,
                'start_bit': start_bit,
                'size': sig_len,
                'byte_order': byte_order,
                'signed': signed,
                'factor': factor,
                'offset': offset_val,
                'min': min_v,
                'max': max_v,
                'unit': str(unit) if unit else '',
                'receivers': receivers,
                'val_desc': val_desc,
                'comment': str(sig_desc) if sig_desc else '',
                'init_raw': init_raw,
                'invalid_raw': invalid_raw,
            })

    # ── Generate DBC text ──
    lines = []
    lines.append('VERSION ""')
    lines.append('')
    lines.append('')
    lines.append('NS_ :')
    lines.append('\tNS_DESC_')
    lines.append('\tCM_')
    lines.append('\tBA_DEF_')
    lines.append('\tBA_')
    lines.append('\tVAL_')
    lines.append('\tCAT_DEF_')
    lines.append('\tCAT_')
    lines.append('\tFILTER')
    lines.append('\tBA_DEF_DEF_')
    lines.append('\tEV_DATA_')
    lines.append('\tENVVAR_DATA_')
    lines.append('\tSGTYPE_')
    lines.append('\tSGTYPE_VAL_')
    lines.append('\tBA_DEF_DEF_REL_')
    lines.append('\tBA_DEF_REL_')
    lines.append('\tBA_DEF_SGTYPE_')
    lines.append('\tBA_DEF_REL_')
    lines.append('\tBA_DEF_DEF_REL_')
    lines.append('\tBU_SGTYPE_')
    lines.append('\tBA_DEF_')
    lines.append('')
    lines.append('')

    # Nodes
    sorted_nodes = sorted(nodes_set)
    lines.append(f'BU_: {" ".join(sorted_nodes)}')
    lines.append('')

    # Messages and signals
    cm_lines = []  # comments
    val_lines = []  # value descriptions
    ba_cycle_lines = []
    ba_sendtype_lines = []
    ba_startval_lines = []
    ba_invalid_lines = []
    ba_nm_lines = []

    for msg_id in sorted(messages.keys()):
        msg = messages[msg_id]
        # DBC name must be max 32 chars, no special chars
        dbc_name = re.sub(r'[^a-zA-Z0-9_]', '_', msg['name'])[:32]
        if not dbc_name[0].isalpha() and dbc_name[0] != '_':
            dbc_name = 'M_' + dbc_name

        lines.append(f'BO_ {msg_id} {dbc_name}: {msg["dlc"]} {msg["transmitter"]}')

        for sig in msg['signals']:
            # DBC signal name max 32 chars
            sig_name = re.sub(r'[^a-zA-Z0-9_]', '_', sig['name'])[:32]
            if not sig_name[0].isalpha() and sig_name[0] != '_':
                sig_name = 'S_' + sig_name

            signed_char = '-' if sig['signed'] else '+'
            # Format factor,offset
            factor_str = f'{sig["factor"]}'
            offset_str = f'{sig["offset"]}'
            if sig['offset'] == 0:
                factor_offset = factor_str
            else:
                factor_offset = f'{factor_str},{offset_str}'

            # Min, max
            min_str = f'{sig["min"]}' if sig['min'] != 0 else '0'
            max_str = f'{sig["max"]}' if sig['max'] != 0 else '0'

            # Receivers or 'Vector__XXX'
            receivers_str = ','.join(sig['receivers']) if sig['receivers'] else 'Vector__XXX'

            lines.append(
                f' SG_ {sig_name} : {sig["start_bit"]}|{sig["size"]}@{sig["byte_order"]}{signed_char}'
                f' ({factor_offset}) [{min_str}|{max_str}] "{sig["unit"]}"  {receivers_str}'
            )

            # Collect comment
            if sig['comment']:
                comment_escaped = sig['comment'].replace('"', '""')
                cm_lines.append(f'CM_ SG_ {msg_id} {sig_name} "{comment_escaped}";')

            # Collect value descriptions
            if sig['val_desc']:
                val_parts = ' '.join(
                    f'{v} "{d.replace(chr(34), chr(34)+chr(34))}"'
                    for v, d in sorted(sig['val_desc'].items())
                )
                val_lines.append(f'VAL_ {msg_id} {sig_name} {val_parts} ;')

            # Collect initial value
            if sig['init_raw'] is not None:
                ba_startval_lines.append(
                    f'BA_ "GenSigStartValue" SG_ {msg_id} {sig_name} {sig["init_raw"]};'
                )

            # Collect invalid value
            if sig['invalid_raw'] is not None:
                ba_invalid_lines.append(
                    f'BA_ "GenSigInvalidValue" SG_ {msg_id} {sig_name} {sig["invalid_raw"]};'
                )

        # Cycle time
        if msg['cycle_time'] > 0:
            ba_cycle_lines.append(f'BA_ "GenMsgCycleTime" BO_ {msg_id} {msg["cycle_time"]};')

        # Send type
        if msg['send_type'] > 0:
            ba_sendtype_lines.append(f'BA_ "GenMsgSendType" BO_ {msg_id} {msg["send_type"]};')

        # NM message
        if msg['is_nm']:
            ba_nm_lines.append(f'BA_ "NmMessage" BO_ {msg_id} 1;')

        lines.append('')

    lines.append('')

    # Add comments
    if cm_lines:
        for l in cm_lines:
            lines.append(l)
        lines.append('')

    # Add value descriptions
    if val_lines:
        for l in val_lines:
            lines.append(l)
        lines.append('')

    # Add BA attribute definitions (standard CANdb++ attributes)
    lines.append('BA_DEF_ BO_ "GenMsgCycleTime" INT 0 65536;')
    lines.append('BA_DEF_ BO_ "GenMsgSendType" INT 0 5;')
    lines.append('BA_DEF_ BO_ "NmMessage" INT 0 1;')
    lines.append('BA_DEF_ SG_ "GenSigStartValue" INT -32768 32767;')
    lines.append('BA_DEF_ SG_ "GenSigInvalidValue" INT -32768 32767;')
    lines.append('BA_DEF_DEF_ "GenMsgCycleTime" 0;')
    lines.append('BA_DEF_DEF_ "GenMsgSendType" 0;')
    lines.append('BA_DEF_DEF_ "NmMessage" 0;')
    lines.append('BA_DEF_DEF_ "GenSigStartValue" 0;')
    lines.append('BA_DEF_DEF_ "GenSigInvalidValue" 0;')
    lines.append('')

    # Add attribute values
    for l in ba_cycle_lines:
        lines.append(l)
    for l in ba_sendtype_lines:
        lines.append(l)
    for l in ba_nm_lines:
        lines.append(l)
    for l in ba_startval_lines:
        lines.append(l)
    for l in ba_invalid_lines:
        lines.append(l)

    lines.append('')

    # Write output
    dbc_text = '\n'.join(lines)
    with open(output_path, 'w', encoding='gbk') as f:
        f.write(dbc_text)

    print(f"Generated DBC: {output_path}")
    print(f"  Messages: {len(messages)}")
    print(f"  Signals: {sum(len(m['signals']) for m in messages.values())}")
    print(f"  Nodes: {len(sorted_nodes)} ({', '.join(sorted_nodes)})")

    # Verify by re-parsing
    from dbc_parser import DbcParser
    parser = DbcParser()
    db = parser.parse(dbc_text)
    print(f"\nVerification: parsed {len(db.messages)} messages, {len(db.nodes)} nodes")
    total_sigs = sum(len(m.signals) for m in db.messages.values())
    print(f"  Total signals: {total_sigs}")
    for msg_id in sorted(db.messages.keys()):
        msg = db.messages[msg_id]
        print(f"  0x{msg_id:03X}: {msg.name} ({len(msg.signals)} signals, DLC={msg.dlc})")

    return db


if __name__ == '__main__':
    build_dbc(EXCEL_PATH, OUTPUT_DBC)
