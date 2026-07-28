#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
signal2dbc.py — 从通信矩阵 Excel 生成 CANoe DBC 文件

支持两种输入格式:
  格式 A (完整矩阵): 包含 Msg ID / Signal Name / Start Byte / Start Bit / Byte Order 等列
  格式 B (简化矩阵): 仅有中文描述, 使用内置帧模板 (北汽智造座椅按摩系统)

用法:
    python signal2dbc.py 通信矩阵.xlsx              # 自动检测格式, 生成 .dbc
    python signal2dbc.py 通信矩阵.xlsx -o my.dbc     # 指定输出文件名
    python signal2dbc.py --list                      # 列出内置帧模板
"""

import sys, os, re, argparse

# ── 依赖检查 ──
try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl 库, 请执行: pip install openpyxl")
    sys.exit(1)


# ================================================================
# 格式 B 的内置帧模板 (北汽智造座椅按摩控制器)
# ================================================================

FRAME_TEMPLATES = [
    # (id_hex, name, sender, receiver, cycle_ms, dlc, signals)
    # signals: (name, start_bit, length, factor, offset, min, max, unit, desc)

    # ── ECU -> dSPACE ──
    {'id_hex': 0x100, 'name': 'ECU_Actuator1_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Act1_StartStop',     7, 1, 1, 0, 0, 1, '', 'Act1 start/stop (0=stop,1=run)'),
        ('Act2_StartStop',     6, 1, 1, 0, 0, 1, '', 'Act2 start/stop'),
        ('Act3_StartStop',     5, 1, 1, 0, 0, 1, '', 'Act3 start/stop'),
        ('Act4_StartStop',     4, 1, 1, 0, 0, 1, '', 'Act4 start/stop'),
        ('Act5_StartStop',     3, 1, 1, 0, 0, 1, '', 'Act5 start/stop'),
        ('Act1_Direction',    15, 1, 1, 0, 0, 1, '', 'Act1 dir (0=CW,1=CCW)'),
        ('Act2_Direction',    14, 1, 1, 0, 0, 1, '', 'Act2 dir'),
        ('Act3_Direction',    13, 1, 1, 0, 0, 1, '', 'Act3 dir'),
        ('Act4_Direction',    12, 1, 1, 0, 0, 1, '', 'Act4 dir'),
        ('Act5_Direction',    11, 1, 1, 0, 0, 1, '', 'Act5 dir'),
        ('Act1_Stroke',       23, 8, 1, 0, 0, 100, '%', 'Act1 stroke 0-100%'),
        ('Act2_Stroke',       31, 8, 1, 0, 0, 100, '%', 'Act2 stroke 0-100%'),
        ('Act3_Stroke',       39, 8, 1, 0, 0, 100, '%', 'Act3 stroke 0-100%'),
        ('Act4_Stroke',       47, 8, 1, 0, 0, 100, '%', 'Act4 stroke 0-100%'),
        ('Act5_Stroke',       55, 8, 1, 0, 0, 100, '%', 'Act5 stroke 0-100%'),
        ('Act_Heartbeat',     63, 8, 1, 0, 0, 255, '', 'Heartbeat 0-255 cycle'),
     ]},
    {'id_hex': 0x101, 'name': 'ECU_MultiActuator_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('MultiAct1_StartStop',  7, 1, 1, 0, 0, 1, '', 'Multi-act1 start/stop'),
        ('MultiAct2_StartStop',  6, 1, 1, 0, 0, 1, '', 'Multi-act2 start/stop'),
        ('MultiAct3_StartStop',  5, 1, 1, 0, 0, 1, '', 'Multi-act3 start/stop'),
        ('MultiAct1_Direction', 15, 1, 1, 0, 0, 1, '', 'Multi-act1 dir'),
        ('MultiAct2_Direction', 14, 1, 1, 0, 0, 1, '', 'Multi-act2 dir'),
        ('MultiAct3_Direction', 13, 1, 1, 0, 0, 1, '', 'Multi-act3 dir'),
        ('MultiAct1_Stroke',    23, 8, 1, 0, 0, 100, '%', 'Multi-act1 stroke'),
        ('MultiAct2_Stroke',    31, 8, 1, 0, 0, 100, '%', 'Multi-act2 stroke'),
        ('MultiAct3_Stroke',    39, 8, 1, 0, 0, 100, '%', 'Multi-act3 stroke'),
        ('MultiAct_Heartbeat',  63, 8, 1, 0, 0, 255, '', 'Heartbeat 0-255'),
     ]},
    {'id_hex': 0x102, 'name': 'ECU_Analog_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('SteeringWheel_Angle',   7, 8, 6, -540, -540, 540, 'deg', 'Steering angle -540~540'),
        ('Gear_Position',        15, 4, 1, 0, 0, 3, '', 'Gear (0=P,1=D,2=R,3=N)'),
        ('Game_Entertain_Info',  23, 8, 1, 0, 0, 255, '', 'Game/entertainment info'),
        ('Airbag_Status',        31, 8, 1, 0, 0, 255, '', 'Airbag working status'),
     ]},
    {'id_hex': 0x103, 'name': 'ECU_Sensor_Data',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Accel_Sensor',           7, 8, 1, 0, 0, 255, '', 'Accelerometer'),
        ('Capacitive_Pressure',   15, 8, 1, 0, 0, 255, '', 'Capacitive pressure'),
        ('AirCell1_Pressure',     23, 8, 1, 0, 0, 255, '', 'Air cell 1 pressure'),
        ('AirCell2_Pressure',     31, 8, 1, 0, 0, 255, '', 'Air cell 2 pressure'),
        ('AirCell3_Pressure',     39, 8, 1, 0, 0, 255, '', 'Air cell 3 pressure'),
        ('AirCell4_Pressure',     47, 8, 1, 0, 0, 255, '', 'Air cell 4 pressure'),
        ('AirCell5_Pressure',     55, 8, 1, 0, 0, 255, '', 'Air cell 5 pressure'),
        ('AirCell6_Pressure',     63, 8, 1, 0, 0, 255, '', 'Air cell 6 pressure'),
     ]},
    {'id_hex': 0x104, 'name': 'ECU_System_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 1000, 'dlc': 8,
     'signals': [
        ('App_Response_Status',  7, 1, 1, 0, 0, 1, '', 'Response status Bit0'),
        ('Func_Enable_Bit0',     6, 1, 1, 0, 0, 1, '', 'Function enable Bit0'),
        ('Func_Enable_Bit1',     5, 1, 1, 0, 0, 1, '', 'Function enable Bit1'),
        ('Func_Group_Select',   15, 8, 1, 0, 0, 6, '', 'Func group select 1-6'),
     ]},
    {'id_hex': 0x105, 'name': 'ECU_Massage_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 1000, 'dlc': 8,
     'signals': [
        ('AutoMassage_Status',    7, 8, 1, 0, 0, 255, '', 'Auto massage status'),
        ('BackMassage_Status',   15, 8, 1, 0, 0, 255, '', 'Back massage status'),
        ('TCM_Massage_Status',   23, 8, 1, 0, 0, 255, '', 'TCM massage status'),
        ('HotCompress_Status',   31, 8, 1, 0, 0, 255, '', 'Hot compress status'),
     ]},
    {'id_hex': 0x106, 'name': 'ECU_Media_Status',
     'sender': 'ECU', 'receiver': 'dSPACE', 'cycle_ms': 500, 'dlc': 8,
     'signals': [
        ('Media_Event_Signal',   7, 8, 1, 0, 0, 255, '', 'Media/event signal'),
     ]},
    # ── dSPACE -> ECU ──
    {'id_hex': 0x200, 'name': 'dSPACE_Actuator1_Cmd',
     'sender': 'dSPACE', 'receiver': 'ECU', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Cmd_Act1_StartStop',     7, 1, 1, 0, 0, 1, '', 'Cmd Act1 start/stop'),
        ('Cmd_Act2_StartStop',     6, 1, 1, 0, 0, 1, '', 'Cmd Act2 start/stop'),
        ('Cmd_Act3_StartStop',     5, 1, 1, 0, 0, 1, '', 'Cmd Act3 start/stop'),
        ('Cmd_Act4_StartStop',     4, 1, 1, 0, 0, 1, '', 'Cmd Act4 start/stop'),
        ('Cmd_Act5_StartStop',     3, 1, 1, 0, 0, 1, '', 'Cmd Act5 start/stop'),
        ('Cmd_Act1_Direction',    15, 1, 1, 0, 0, 1, '', 'Cmd Act1 dir'),
        ('Cmd_Act2_Direction',    14, 1, 1, 0, 0, 1, '', 'Cmd Act2 dir'),
        ('Cmd_Act3_Direction',    13, 1, 1, 0, 0, 1, '', 'Cmd Act3 dir'),
        ('Cmd_Act4_Direction',    12, 1, 1, 0, 0, 1, '', 'Cmd Act4 dir'),
        ('Cmd_Act5_Direction',    11, 1, 1, 0, 0, 1, '', 'Cmd Act5 dir'),
        ('Cmd_Act1_Stroke',       23, 8, 1, 0, 0, 100, '%', 'Cmd Act1 stroke'),
        ('Cmd_Act2_Stroke',       31, 8, 1, 0, 0, 100, '%', 'Cmd Act2 stroke'),
        ('Cmd_Act3_Stroke',       39, 8, 1, 0, 0, 100, '%', 'Cmd Act3 stroke'),
        ('Cmd_Act4_Stroke',       47, 8, 1, 0, 0, 100, '%', 'Cmd Act4 stroke'),
        ('Cmd_Act5_Stroke',       55, 8, 1, 0, 0, 100, '%', 'Cmd Act5 stroke'),
     ]},
    {'id_hex': 0x201, 'name': 'dSPACE_MultiActuator_Cmd',
     'sender': 'dSPACE', 'receiver': 'ECU', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Cmd_MultiAct1_StartStop',  7, 1, 1, 0, 0, 1, '', 'Cmd multi-act1 start/stop'),
        ('Cmd_MultiAct2_StartStop',  6, 1, 1, 0, 0, 1, '', 'Cmd multi-act2 start/stop'),
        ('Cmd_MultiAct3_StartStop',  5, 1, 1, 0, 0, 1, '', 'Cmd multi-act3 start/stop'),
        ('Cmd_MultiAct1_Direction', 15, 1, 1, 0, 0, 1, '', 'Cmd multi-act1 dir'),
        ('Cmd_MultiAct2_Direction', 14, 1, 1, 0, 0, 1, '', 'Cmd multi-act2 dir'),
        ('Cmd_MultiAct3_Direction', 13, 1, 1, 0, 0, 1, '', 'Cmd multi-act3 dir'),
        ('Cmd_MultiAct1_Stroke',    23, 8, 1, 0, 0, 100, '%', 'Cmd multi-act1 stroke'),
        ('Cmd_MultiAct2_Stroke',    31, 8, 1, 0, 0, 100, '%', 'Cmd multi-act2 stroke'),
        ('Cmd_MultiAct3_Stroke',    39, 8, 1, 0, 0, 100, '%', 'Cmd multi-act3 stroke'),
     ]},
    {'id_hex': 0x202, 'name': 'dSPACE_AirCell_Cmd',
     'sender': 'dSPACE', 'receiver': 'ECU', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Cmd_AirCell1',    7, 8, 1, 0, 0, 255, '', 'Cmd air cell 1'),
        ('Cmd_AirCell2',   15, 8, 1, 0, 0, 255, '', 'Cmd air cell 2'),
        ('Cmd_AirCell3',   23, 8, 1, 0, 0, 255, '', 'Cmd air cell 3'),
        ('Cmd_AirCell4',   31, 8, 1, 0, 0, 255, '', 'Cmd air cell 4'),
        ('Cmd_AirCell5',   39, 8, 1, 0, 0, 255, '', 'Cmd air cell 5'),
        ('Cmd_AirCell6',   47, 8, 1, 0, 0, 255, '', 'Cmd air cell 6'),
        ('Cmd_AirCell7',   55, 8, 1, 0, 0, 255, '', 'Cmd air cell 7'),
        ('Cmd_AirCell8',   63, 8, 1, 0, 0, 255, '', 'Cmd air cell 8'),
     ]},
    {'id_hex': 0x203, 'name': 'dSPACE_Massage_Cmd',
     'sender': 'dSPACE', 'receiver': 'ECU', 'cycle_ms': 1000, 'dlc': 8,
     'signals': [
        ('Cmd_AutoMassage',     7, 8, 1, 0, 0, 255, '', 'Cmd auto massage'),
        ('Cmd_BackMassage',    15, 8, 1, 0, 0, 255, '', 'Cmd back massage'),
        ('Cmd_TCM_Massage',    23, 8, 1, 0, 0, 255, '', 'Cmd TCM massage'),
        ('Cmd_HotCompress',    31, 8, 1, 0, 0, 255, '', 'Cmd hot compress'),
     ]},
    {'id_hex': 0x204, 'name': 'dSPACE_Misc_Cmd',
     'sender': 'dSPACE', 'receiver': 'ECU', 'cycle_ms': 100, 'dlc': 8,
     'signals': [
        ('Cmd_Sensor_Config',   7, 8, 1, 0, 0, 255, '', 'Cmd sensor config'),
        ('Cmd_Game_Control',   15, 8, 1, 0, 0, 255, '', 'Cmd game control'),
        ('Cmd_Airbag_Work',    23, 8, 1, 0, 0, 255, '', 'Cmd airbag work'),
     ]},
]


# ================================================================
# 动态解析: 从标准矩阵格式的 Excel 中提取帧和信号
# ================================================================

def detect_matrix_format(ws):
    """检测 Matrix sheet 是否为标准格式 (含 Msg ID / Signal Name / Start Bit 等列)"""
    if ws.max_row < 3 or ws.max_column < 15:
        return False
    header = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or '').lower().replace('\n', ' ').strip()
        header[c] = v

    # 检查必需列
    needed = ['msg id', 'signal name', 'byte order', 'start bit', 'bit length']
    found = [any(kw in h for h in header.values()) for kw in needed]
    return all(found)


def parse_dynamic(ws):
    """从标准矩阵格式动态解析帧和信号"""
    # 定位列索引
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or '').lower().replace('\n', ' ').strip()
        if 'msg name' in v or 'message name' in v:       col_map['msg_name'] = c
        elif 'msg id' in v or 'message id' in v:           col_map['msg_id'] = c
        elif 'msg cycle' in v or 'cycle time' in v:        col_map['cycle'] = c
        elif 'msg length' in v or 'msg len' in v:          col_map['dlc'] = c
        elif 'signal name' in v:                           col_map['sig_name'] = c
        elif 'signal value' in v or 'value description' in v: col_map['val_desc'] = c
        elif 'byte order' in v or '格式' in v:             col_map['byte_order'] = c
        elif 'start byte' in v:                            col_map['start_byte'] = c
        elif 'start bit' in v:                             col_map['start_bit'] = c
        elif 'bit length' in v or 'signal length' in v:    col_map['bit_len'] = c
        elif 'data type' in v or 'signal type' in v:       col_map['data_type'] = c
        elif 'resolution' in v or '因子' in v or 'factor' in v: col_map['factor'] = c
        elif 'offset' in v and 'offset' not in col_map:    col_map['offset'] = c
        elif 'signal min' in v or '最小值' in v:            col_map['min'] = c
        elif 'signal max' in v or '最大值' in v:            col_map['max'] = c
        elif 'initial' in v or '初始' in v:                col_map['init'] = c
        elif 'invalid' in v or '无效' in v:                col_map['invalid'] = c
        elif 'unit' in v:                                  col_map['unit'] = c

    # 扫描节点行 (第2行可能是节点标题)
    node_names = []
    node_cols = {}
    if ws.max_row >= 2:
        for c in range(1, ws.max_column + 1):
            v = ws.cell(2, c).value
            if v and isinstance(v, str) and re.match(r'^[A-Z][A-Z0-9_-]{1,15}$', v.strip()):
                node_names.append(v.strip())
                node_cols[v.strip()] = c

    frames = []
    current = None

    for r in range(3, ws.max_row + 1):
        msg_name = ws.cell(r, col_map['msg_name']).value if 'msg_name' in col_map else None
        sig_name = ws.cell(r, col_map['sig_name']).value if 'sig_name' in col_map else None

        if msg_name and str(msg_name).strip():
            # 新消息开始
            if current and current['signals']:
                frames.append(current)

            # 解析 ID
            id_raw = ws.cell(r, col_map['msg_id']).value if 'msg_id' in col_map else None
            can_id = 0
            if id_raw:
                id_str = str(id_raw).strip()
                if id_str.startswith('0x') or id_str.startswith('0X'):
                    can_id = int(id_str, 16)
                elif id_str.isdigit():
                    can_id = int(id_str)

            # 解析周期 / DLC
            cycle = ws.cell(r, col_map['cycle']).value if 'cycle' in col_map else None
            dlc = ws.cell(r, col_map['dlc']).value if 'dlc' in col_map else None

            # 找发送/接收节点
            sender = None
            receivers = []
            for nn, nc in node_cols.items():
                v = ws.cell(r, nc).value
                if v and str(v).strip().upper() == 'S':
                    sender = nn
                elif v and str(v).strip().upper() == 'R':
                    receivers.append(nn)

            # 如果没有节点列标记, 从消息名推断发送方
            if not sender:
                msg_prefix = msg_name.split('_')[0] if '_' in str(msg_name) else str(msg_name)
                for nn in node_names:
                    if nn.replace('-', '_').startswith(msg_prefix):
                        sender = nn
                        break

            current = {
                'name': str(msg_name).strip(),
                'id': can_id,
                'cycle': int(cycle) if cycle else 100,
                'dlc': int(dlc) if dlc else 8,
                'sender': sender or 'Node_A',
                'receiver': receivers[0] if receivers else 'Node_B',
                'all_receivers': receivers,
                'signals': [],
                'val_defs_all': {},
            }
            # fallback: 如果有多个接收者, 用第一个
            if not receivers:
                current['receiver'] = 'Node_B'

        if sig_name and current:
            sig_name_str = str(sig_name).strip()
            # 过滤非法字符 -> 下划线
            safe_name = re.sub(r'[^A-Za-z0-9_]', '_', sig_name_str)
            safe_name = re.sub(r'_+', '_', safe_name).strip('_')

            start_byte = int(ws.cell(r, col_map['start_byte']).value) if 'start_byte' in col_map and ws.cell(r, col_map['start_byte']).value is not None else 0
            start_bit_raw = ws.cell(r, col_map['start_bit']).value if 'start_bit' in col_map else 0
            start_bit = int(start_bit_raw) if start_bit_raw is not None else 0
            bit_len = int(ws.cell(r, col_map['bit_len']).value) if 'bit_len' in col_map and ws.cell(r, col_map['bit_len']).value is not None else 8
            byte_order = str(ws.cell(r, col_map['byte_order']).value or 'Motorola').strip() if 'byte_order' in col_map else 'Motorola'
            factor = float(ws.cell(r, col_map['factor']).value) if 'factor' in col_map and ws.cell(r, col_map['factor']).value is not None else 1
            offset = float(ws.cell(r, col_map['offset']).value) if 'offset' in col_map and ws.cell(r, col_map['offset']).value is not None else 0
            pmin = float(ws.cell(r, col_map['min']).value) if 'min' in col_map and ws.cell(r, col_map['min']).value is not None else 0
            pmax = float(ws.cell(r, col_map['max']).value) if 'max' in col_map and ws.cell(r, col_map['max']).value is not None else 255
            unit = str(ws.cell(r, col_map['unit']).value or '').strip() if 'unit' in col_map else ''

            # Excel "Start Bit" 就是信号的 LSB 位置
            # 字节序按矩阵要求用 Motorola
            dbc_start = start_bit
            dbc_byte_order = '@1+'  # Motorola, Unsigned

            # 解析枚举值
            val_desc = ws.cell(r, col_map['val_desc']).value if 'val_desc' in col_map else None
            val_defs = {}
            if val_desc:
                for line in str(val_desc).split('\n'):
                    line = line.strip()
                    if ':' in line:
                        parts = line.split(':', 1)
                        val_str = parts[0].strip()
                        desc_str = parts[1].strip()[:60]
                        if val_str.startswith('0x'):
                            try:
                                val_defs[int(val_str, 16)] = desc_str
                            except:
                                pass
                        elif val_str.isdigit():
                            val_defs[int(val_str)] = desc_str

            if val_defs:
                current['val_defs_all'][safe_name] = val_defs

            desc = f'{sig_name_str}'
            current['signals'].append((safe_name, dbc_start, bit_len, factor, offset, pmin, pmax, unit, desc, dbc_byte_order))

    if current and current['signals']:
        frames.append(current)

    return frames


# ================================================================
# DBC 文本生成器
# ================================================================

def sanitize_node(name):
    """将节点名中的非法字符替换为下划线 (DBC 只允许 A-Za-z0-9_)"""
    return re.sub(r'[^A-Za-z0-9_]', '_', name)


def gen_dbc(frames):
    """根据帧列表生成 DBC 文本 (纯 ASCII)"""
    lines = []

    lines.append('VERSION ""')
    lines.append('')

    lines.append('NS_ :')
    for ns in ['NS_DESC_', 'CM_', 'BA_DEF_', 'BA_DEF_DEF_',
               'BA_DEF_DEF_REL_', 'BA_DEF_SGTYPE_', 'BA_DEF_REL_',
               'BA_DEF_DEF_', 'BU_DEF_REL_', 'BA_', 'BA_DEF_',
               'VAL_', 'CAT_DEF_', 'CAT_', 'FILTER',
               'BA_DEF_DEF_', 'EV_DATA_', 'ENVVAR_DATA_', 'SGTYPE_',
               'SGTYPE_VAL_', 'BA_DEF_SGTYPE_', 'BA_SGTYPE_', 'SIG_VALTYPE_',
               'SIGTYPE_VALTYPE_', 'BO_TX_BU_', 'BA_DEF_REL_',
               'BA_DEF_DEF_REL_', 'BA_REL_', 'BA_DEF_DEF_', 'BU_BO_REL_',
               'BA_DEF_', 'BU_SGP_REL_', 'BU_SGT_REL_', 'SG_MUL_VAL_']:
        lines.append(f'  {ns}')
    lines.append('')

    lines.append('BS_:')
    lines.append('')

    all_nodes = set()
    for f in frames:
        if f['sender']:   all_nodes.add(sanitize_node(f['sender']))
        if f['receiver']: all_nodes.add(sanitize_node(f['receiver']))
        if 'all_receivers' in f:
            for rc in f['all_receivers']:
                all_nodes.add(sanitize_node(rc))
    lines.append(f'BU_: {" ".join(sorted(all_nodes))}')
    lines.append('')

    # BO_ + SG_
    for f in frames:
        sigs = []
        for sig in f['signals']:
            # 支持第10个元素为字节序标记，不兼容旧格式时默认 @1+ (Motorola)
            if len(sig) >= 10:
                name, start, length, factor, offset, pmin, pmax, unit, desc, bo = sig
            else:
                name, start, length, factor, offset, pmin, pmax, unit, desc = sig
                bo = '@1+'
            rcv = sanitize_node(f['receiver'])
            min_v = pmin
            max_v = pmax if pmax > pmin else pmin + (1 << length) - 1
            sigs.append(
                f'  SG_ {name} : {start}|{length}{bo} ({factor:g},{offset:g}) [{min_v:g}|{max_v:g}] "{unit}"  {rcv}'
            )
        lines.append(f'BO_ {f["id"]} {f["name"]}: {f["dlc"]} {sanitize_node(f["sender"])}')
        lines.extend(sigs)
        lines.append('')

    # CM_ 信号注释
    for f in frames:
        for sig in f['signals']:
            name = sig[0]
            desc = sig[8] if len(sig) >= 9 else (sig[-1] if len(sig) >= 9 else '')
            if desc:
                lines.append(f'CM_ SG_ {f["id"]} {name} "{desc}";')
    lines.append('')

    # CM_ 帧注释
    for f in frames:
        cmt = f.get('comment', f['name'])
        lines.append(f'CM_ BO_ {f["id"]} "{cmt}";')
    lines.append('')

    # BA_DEF_
    lines.append('BA_DEF_ BO_ "GenMsgCycleTime" INT 0 65535;')
    lines.append('BA_DEF_ BO_ "GenMsgSendType" STRING;')
    lines.append('')
    lines.append('BA_DEF_DEF_ "GenMsgCycleTime" 0;')
    lines.append('BA_DEF_DEF_ "GenMsgSendType" "Cyclic";')
    lines.append('')

    for f in frames:
        lines.append(f'BA_ "GenMsgCycleTime" BO_ {f["id"]} {f["cycle"]};')
    lines.append('')

    # VAL_ 枚举
    for f in frames:
        fid = f['id']
        val_defs = f.get('val_defs_all', {})
        # 也检查 signals 中的描述推导枚举
        for sig in f['signals']:
            name = sig[0]
            bit_len = sig[2]
            if name in val_defs:
                for k, v in sorted(val_defs[name].items()):
                    lines.append(f'VAL_ {fid} {name} {k} "{v}" ;')
            elif bit_len == 1:
                lines.append(f'VAL_ {fid} {name} 1 "Active" 0 "Inactive" ;')
            elif bit_len <= 2 and name.endswith('_Direction'):
                lines.append(f'VAL_ {fid} {name} 1 "Forward" 0 "CW" ;')

    lines.append('')
    return '\n'.join(lines)


def save_dbc(dbc_text, output_path):
    """ANSI 编码写入（自动清除非 Latin-1 字符）"""
    # 替换所有非 Latin-1 字符为 ASCII 近似或删除
    cleaned = []
    for ch in dbc_text:
        if ord(ch) < 128:
            cleaned.append(ch)
        elif ord(ch) < 256:
            # Latin-1 补充字符 (ÅÄÖ等) 保留
            cleaned.append(ch)
        else:
            # 中文等非 Latin-1 字符替换为空格
            cleaned.append(' ')
    dbc_clean = ''.join(cleaned)
    # 合并连续空格
    dbc_clean = re.sub(r'  +', ' ', dbc_clean)
    with open(output_path, 'w', encoding='windows-1252') as f:
        f.write(dbc_clean)


def print_summary(frames, output_path):
    total_sigs = sum(len(f['signals']) for f in frames)
    print(f'DBC 文件: {output_path}')
    print(f'帧数: {len(frames)}')
    print(f'信号数: {total_sigs}')
    print()
    print(f'{"ID":>8s}  {"Frame Name":35s} {"DLC":4s} {"Cycle":6s}  {"Direction":20s}  {"Signals":8s}')
    print('-' * 90)
    for f in frames:
        direction = f'{f["sender"]:8s} -> {f["receiver"]:8s}'
        print(f'  0x{f["id"]:03X}  {f["name"]:35s} {f["dlc"]:4d} {f["cycle"]:5d}ms  {direction:20s}  {len(f["signals"]):3d}')
    print()


# ================================================================
# 主入口
# ================================================================

def main():
    parser = argparse.ArgumentParser(
        description='从通信矩阵 Excel 生成 CANoe DBC 文件',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument('input', nargs='?', help='Excel 信号矩阵文件路径')
    parser.add_argument('-o', '--output', help='输出 DBC 文件名')
    parser.add_argument('--list', action='store_true', help='列出内置帧模板')
    args = parser.parse_args()

    # 确定输入路径
    if args.input:
        input_path = args.input
        if args.output:
            output_path = args.output
        else:
            base, _ = os.path.splitext(input_path)
            output_path = base + '.dbc'
    else:
        xlsx_files = [f for f in os.listdir('.') if f.lower().endswith('.xlsx') and not f.startswith('~')]
        if not xlsx_files:
            print('错误: 未指定输入文件，且当前目录下没有 .xlsx 文件')
            parser.print_help()
            sys.exit(1)
        input_path = xlsx_files[0]
        base, _ = os.path.splitext(input_path)
        output_path = base + '.dbc'
        print(f'自动选择: {input_path}')

    if args.list:
        print_summary(FRAME_TEMPLATES, output_path)
        return

    print(f'输入: {input_path}')
    print(f'输出: {output_path}')

    wb = openpyxl.load_workbook(input_path)
    ws = None

    # 优先用 Matrix sheet, 其次第一个
    if 'Matrix' in wb.sheetnames:
        ws = wb['Matrix']
    elif 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    else:
        ws = wb[wb.sheetnames[0]]

    # 检测格式
    if detect_matrix_format(ws):
        print('检测到标准矩阵格式 → 动态解析')
        frames = parse_dynamic(ws)
    else:
        print('检测到简化格式 → 使用内置帧模板')
        # 转换为统一格式
        frames = []
        for ft in FRAME_TEMPLATES:
            frames.append({
                'id': ft['id_hex'],
                'name': ft['name'],
                'sender': ft['sender'],
                'receiver': ft['receiver'],
                'cycle': ft['cycle_ms'],
                'dlc': ft['dlc'],
                'signals': ft['signals'],
                'val_defs_all': {},
                'comment': None,
            })

    if not frames:
        print('错误: 没有解析到任何帧')
        sys.exit(1)

    dbc_text = gen_dbc(frames)
    save_dbc(dbc_text, output_path)
    print_summary(frames, output_path)
    print('生成完成!')


if __name__ == '__main__':
    main()
